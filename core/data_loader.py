import json
import logging
import os
import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd

_log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Storage-format helpers
# ---------------------------------------------------------------------------

def _find_project_root(path: Path) -> Path | None:
    """Walk up from path to find the directory containing project.json."""
    current = path if path.is_dir() else path.parent
    for _ in range(10):
        if (current / "project.json").exists():
            return current
        parent = current.parent
        if parent == current:
            break
        current = parent
    return None


def get_storage_format(project_path: Path) -> str:
    """Return 'parquet' or 'csv' for the given project. Defaults to 'csv' for old projects."""
    pj = Path(project_path) / "project.json"
    if pj.exists():
        try:
            with open(pj, encoding="utf-8") as f:
                return json.load(f).get("storage_format", "csv")
        except Exception:
            pass
    return "csv"


def read_table(file_path: Path) -> pd.DataFrame:
    """Read a tabular file (CSV or Parquet) into a DataFrame.

    Tries the given path first; if not found, tries the alternate extension.
    All columns are returned as strings (matching the CSV behaviour).
    """
    file_path = Path(file_path)

    def _read(p: Path) -> pd.DataFrame:
        if p.suffix == ".parquet":
            try:
                df = pd.read_parquet(p)
                return df.astype(str).where(df.notna(), "")
            except Exception as e:
                msg = str(e).lower()
                if "schema" in msg or "invalid" in msg or "arrow" in msg:
                    raise ValueError(
                        f"The data file '{p.name}' has an unexpected format.\n"
                        "The project may need to be re-imported."
                    ) from e
                raise OSError(f"Could not read data file '{p.name}': {e}") from e
        return pd.read_csv(p, dtype=str, keep_default_na=False, encoding="utf-8")

    if file_path.exists():
        return _read(file_path)
    alt = file_path.with_suffix(".csv" if file_path.suffix == ".parquet" else ".parquet")
    if alt.exists():
        return _read(alt)
    raise FileNotFoundError(f"Table file not found: '{file_path}'")


def write_table(df: pd.DataFrame, dest_path: Path) -> Path:
    """Write a DataFrame to disk as CSV or Parquet based on dest_path's extension.

    Raises FileExistsError if dest_path is .parquet but a .csv file already
    exists at the same stem — existing CSV projects are never silently converted.
    Returns the path actually written.
    """
    dest_path = Path(dest_path)
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = dest_path.with_suffix(dest_path.suffix + ".tmp")
    try:
        if dest_path.suffix == ".parquet":
            csv_sibling = dest_path.with_suffix(".csv")
            if csv_sibling.exists():
                raise FileExistsError(
                    f"Cannot write Parquet: a CSV file already exists at '{csv_sibling}'. "
                    "Existing projects stay in their original format."
                )
            try:
                # Cast every column to plain string before writing so pyarrow
                # never stores date-formatted values as timestamp/date types.
                df_str = df.copy()
                for col in df_str.columns:
                    df_str[col] = df_str[col].fillna("").astype(str)
                df_str.to_parquet(tmp_path, index=False, compression="snappy")
            except OSError as e:
                if getattr(e, "errno", None) == 28 or getattr(e, "winerror", None) == 112:
                    raise OSError(
                        "Not enough disk space to save the file.\n"
                        "Free up space and try again."
                    ) from e
                raise
            # Read-back validation: confirm row count matches
            written = pd.read_parquet(tmp_path)
            if len(written) != len(df_str):
                raise OSError(
                    f"Parquet write validation failed for '{dest_path.name}': "
                    f"wrote {len(df)} rows but read back {len(written)}. "
                    "The file may be corrupted — please try again."
                )
        else:
            try:
                df.to_csv(tmp_path, index=False, encoding="utf-8")
            except OSError as e:
                if getattr(e, "errno", None) == 28 or getattr(e, "winerror", None) == 112:
                    raise OSError(
                        "Not enough disk space to save the file.\n"
                        "Free up space and try again."
                    ) from e
                raise
        # Atomic replace: tmp → destination (safe even if dest already exists)
        os.replace(tmp_path, dest_path)
    except BaseException:
        # Clean up the temp file if anything went wrong
        try:
            tmp_path.unlink(missing_ok=True)
        except OSError:
            pass
        raise
    return dest_path


def _excel_numfmt_to_strftime(number_format: str) -> str | None:
    """Convert an Excel number format string to a Python strftime string.

    Returns None if the format doesn't contain date tokens.
    """
    fmt = (number_format or "").strip()
    # Strip color codes e.g. [Red], [Black]
    fmt = re.sub(r'\[[^\]]+\]', '', fmt)
    # Strip quoted literal text e.g. "dd \"of\" mmm"
    fmt = re.sub(r'"[^"]*"', '', fmt)
    # Strip the text-cell suffix — Excel format strings can contain
    # positive;negative;zero;text sections separated by semicolons.
    # Only the first section applies to date values.
    fmt = fmt.split(';')[0]

    # Must contain at least one unambiguous date/year token to qualify
    if not re.search(r'y{1,4}|d{1,4}|mmm', fmt, re.IGNORECASE):
        return None

    result = fmt
    # Replace tokens longest-first to avoid partial matches.
    # Year / month-name tokens first.
    result = re.sub(r'(?i)yyyy', '%Y', result)
    result = re.sub(r'(?i)yy',   '%y', result)
    result = re.sub(r'(?i)mmmm', '%B', result)
    result = re.sub(r'(?i)mmm',  '%b', result)
    # Day tokens before any single-letter replacements.
    result = re.sub(r'(?i)dddd', '%A', result)
    result = re.sub(r'(?i)ddd',  '%a', result)
    result = re.sub(r'(?i)dd',   '%d', result)
    result = re.sub(r'(?i)d',    '%d', result)
    # Hours + minutes: handle "hh:mm" and "h:mm" BEFORE replacing standalone
    # mm so that minute tokens (context: after hours) map to %M, not %m.
    result = re.sub(r'(?i)hh:mm', '%H:%M', result)
    result = re.sub(r'(?i)h:mm',  '%H:%M', result)
    result = re.sub(r'(?i)hh',    '%H',    result)
    result = re.sub(r'(?i)h',     '%H',    result)
    # Seconds
    result = re.sub(r'(?i)ss', '%S', result)
    # Any remaining mm / m are month tokens (time component already handled).
    result = re.sub(r'(?i)mm', '%m', result)
    result = re.sub(r'(?i)m',  '%m', result)

    return result if '%' in result else None


def _raise_excel_error(file_path: Path, exc: Exception) -> None:
    """Re-raise an Excel open/read error as a typed, user-friendly exception."""
    _log.error("Failed to open Excel file '%s': %s", file_path, exc, exc_info=True)
    msg = str(exc).lower()
    if isinstance(exc, PermissionError):
        raise PermissionError(
            f"'{file_path.name}' is open in another program.\n"
            "Close the file in Excel and try again."
        ) from exc
    if "password" in msg or "encrypt" in msg:
        raise ValueError(
            f"'{file_path.name}' is password-protected.\n"
            "Remove the password in Excel and try again."
        ) from exc
    if "bad zip" in msg or "not a zip" in msg:
        raise ValueError(
            f"'{file_path.name}' appears to be corrupted and cannot be opened.\n"
            "Try re-exporting the file from its source."
        ) from exc
    raise ValueError(
        f"Could not open '{file_path.name}': {exc}"
    ) from exc


def load_excel_sheets(file_path: Path) -> list:
    """Return the list of sheet names in an Excel file."""
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    try:
        xl = pd.ExcelFile(file_path, engine="openpyxl")
        sheets = xl.sheet_names
        _log.info("File loaded: '%s' (%d sheet(s))", file_path.name, len(sheets))
        return sheets
    except (FileNotFoundError, ValueError):
        raise
    except Exception as e:
        _raise_excel_error(file_path, e)


def _find_header_row(ws) -> int:
    """Detect which row contains the column headers in a worksheet.

    Scans the first 20 rows and returns the 1-based index of the first row
    that reaches at least 80% of the maximum non-empty cell count seen.
    This correctly handles files where rows above the header are titles,
    logos, blank spacers, or other decorative content.
    Falls back to row 1 if the sheet is empty or detection is inconclusive.
    """
    scan_limit = min(ws.max_row or 1, 20)
    counts = []
    for r in range(1, scan_limit + 1):
        count = sum(1 for cell in ws[r] if cell.value is not None)
        counts.append(count)

    if not counts or max(counts) == 0:
        return 1

    max_count = max(counts)
    threshold = max_count * 0.8

    for i, count in enumerate(counts):
        if count >= threshold:
            return i + 1  # convert to 1-based row index

    return 1


def detect_merged_cells(file_path: Path, sheet_name: str) -> bool:
    """Return True if *sheet_name* contains any merged cell ranges.

    openpyxl silently reads the top-left value of a merged region and returns
    None for all other cells in the merge — no error, no warning.  Callers use
    this to decide whether to show a user-facing advisory.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(str(Path(file_path)), data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return False
        has_merged = bool(wb[sheet_name].merged_cells.ranges)
        wb.close()
        return has_merged
    except Exception:
        return False


def detect_header_row(file_path: Path, sheet_name: str) -> int:
    """Return the 1-based index of the header row for *sheet_name* in *file_path*."""
    import openpyxl
    file_path = Path(file_path)
    try:
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return 1
        result = _find_header_row(wb[sheet_name])
        wb.close()
        return result
    except Exception:
        return 1


def get_sheet_as_dataframe(file_path: Path, sheet_name: str, header_row: int | None = None) -> pd.DataFrame:
    """Read a single sheet from an Excel file and return a DataFrame.

    Automatically detects the header row (it may not always be row 1 — some
    files have title rows, blank rows, or logos above the actual columns).
    All rows below the detected header row are loaded as data/transaction rows.

    Date/time cells are formatted as strings that match their original Excel
    display format (e.g. dd/mm/yyyy), so the value round-trips correctly
    through CSV and back into the final output workbook.
    """
    import openpyxl

    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    try:
        try:
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
        except PermissionError as e:
            _raise_excel_error(file_path, e)
        except Exception as e:
            _raise_excel_error(file_path, e)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in '{file_path}'")
        ws = wb[sheet_name]

        if not ws.max_row:
            wb.close()
            return pd.DataFrame()

        # --- determine the header row (manual override or auto-detect) ---
        header_row = header_row if header_row is not None else _find_header_row(ws)
        data_start_row = header_row + 1

        # --- headers (detected row) ---
        headers = [
            str(cell.value).strip() if cell.value is not None else f"Col{i}"
            for i, cell in enumerate(ws[header_row], start=1)
        ]

        # --- detect per-column strftime format from first ~20 data rows ---
        col_strftime: dict[int, str] = {}   # 0-based column index → strftime
        scan_end = min(ws.max_row, data_start_row + 19)
        for ci in range(len(headers)):
            for row_cells in ws.iter_rows(min_row=data_start_row, max_row=scan_end,
                                          min_col=ci + 1, max_col=ci + 1):
                for cell in row_cells:
                    if cell.value is not None and isinstance(cell.value, (date, datetime)):
                        fmt = _excel_numfmt_to_strftime(cell.number_format or "")
                        if fmt:
                            col_strftime[ci] = fmt
                        break
                if ci in col_strftime:
                    break

        # --- build data rows (everything below the header) ---
        data_rows = []
        for row_cells in ws.iter_rows(min_row=data_start_row, values_only=False):
            row_data = []
            for ci in range(len(headers)):
                cell = row_cells[ci] if ci < len(row_cells) else None
                val = cell.value if cell is not None else None

                if val is None:
                    row_data.append("")
                elif ci in col_strftime and isinstance(val, (date, datetime)):
                    try:
                        row_data.append(val.strftime(col_strftime[ci]))
                    except Exception:
                        row_data.append(str(val))
                else:
                    row_data.append(str(val))
            data_rows.append(row_data)

        wb.close()
        df = pd.DataFrame(data_rows, columns=headers)
        return df

    except (FileNotFoundError, ValueError):
        raise
    except Exception as e:
        raise ValueError(f"Failed to read sheet '{sheet_name}' from '{file_path}': {e}") from e


def get_sheets_as_dataframes(
    file_path: Path,
    sheets: list[tuple[str, int | None]],
) -> dict[str, pd.DataFrame]:
    """Read multiple sheets from one Excel file in a single workbook open.

    sheets: list of (sheet_name, header_row). header_row=None triggers auto-detect.
    Returns {sheet_name: DataFrame}.  Raises on any missing sheet or I/O error.
    """
    import openpyxl

    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    try:
        wb = openpyxl.load_workbook(str(file_path), data_only=True)
    except PermissionError as e:
        _raise_excel_error(file_path, e)
    except Exception as e:
        _raise_excel_error(file_path, e)
    result: dict[str, pd.DataFrame] = {}
    try:
        for sheet_name, header_row in sheets:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in '{file_path}'")
            ws = wb[sheet_name]
            if not ws.max_row:
                result[sheet_name] = pd.DataFrame()
                continue

            hr = header_row if header_row is not None else _find_header_row(ws)
            data_start = hr + 1
            headers = [
                str(cell.value).strip() if cell.value is not None else f"Col{i}"
                for i, cell in enumerate(ws[hr], start=1)
            ]

            col_strftime: dict[int, str] = {}
            scan_end = min(ws.max_row, data_start + 19)
            for ci in range(len(headers)):
                for row_cells in ws.iter_rows(min_row=data_start, max_row=scan_end,
                                              min_col=ci + 1, max_col=ci + 1):
                    for cell in row_cells:
                        if cell.value is not None and isinstance(cell.value, (date, datetime)):
                            fmt = _excel_numfmt_to_strftime(cell.number_format or "")
                            if fmt:
                                col_strftime[ci] = fmt
                        break
                    if ci in col_strftime:
                        break

            data_rows = []
            for row_cells in ws.iter_rows(min_row=data_start, values_only=False):
                row_data = []
                for ci in range(len(headers)):
                    cell = row_cells[ci] if ci < len(row_cells) else None
                    val = cell.value if cell is not None else None
                    if val is None:
                        row_data.append("")
                    elif ci in col_strftime and isinstance(val, (date, datetime)):
                        try:
                            row_data.append(val.strftime(col_strftime[ci]))
                        except Exception:
                            row_data.append(str(val))
                    else:
                        row_data.append(str(val))
                data_rows.append(row_data)

            result[sheet_name] = pd.DataFrame(data_rows, columns=headers)
    finally:
        wb.close()

    return result


def save_as_csv(df: pd.DataFrame, dest_path: Path) -> None:
    """Save a DataFrame to disk, using Parquet for new projects and CSV for existing ones.

    Callers may pass a .csv path even for Parquet projects — the extension is
    rewritten automatically based on the project's storage_format field.
    """
    dest_path = Path(dest_path)
    root = _find_project_root(dest_path)
    fmt = get_storage_format(root) if root else "csv"
    if fmt == "parquet":
        dest_path = dest_path.with_suffix(".parquet")
    try:
        write_table(df, dest_path)
    except OSError as e:
        raise OSError(f"Failed to save table to '{dest_path}': {e}") from e


def save_as_json(df: pd.DataFrame, dest_path: Path) -> None:
    """Save a DataFrame as a JSON file (records format, one object per row)."""
    dest_path = Path(dest_path)
    try:
        dest_path.parent.mkdir(parents=True, exist_ok=True)
        records = df.to_dict(orient="records")
        with open(dest_path, "w", encoding="utf-8") as f:
            json.dump(records, f, indent=2, ensure_ascii=False)
    except OSError as e:
        raise OSError(f"Failed to save JSON to '{dest_path}': {e}") from e


def load_csv(file_path: Path) -> pd.DataFrame:
    """Load a tabular file as a DataFrame (all columns as strings).

    Accepts .csv or .parquet paths; falls back to the alternate extension if
    the given path does not exist, preserving backward compatibility.
    """
    file_path = Path(file_path)
    try:
        return read_table(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"Table file not found: {file_path}")
    except Exception as e:
        raise ValueError(f"Failed to load table '{file_path}': {e}") from e


def load_dim_json(file_path: Path) -> pd.DataFrame:
    """Load a dim JSON file (records list) as a DataFrame."""
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Dim JSON file not found: {file_path}")
    try:
        with open(file_path, encoding="utf-8") as f:
            records = json.load(f)
        return pd.DataFrame(records).astype(str)
    except Exception as e:
        raise ValueError(f"Failed to load dim JSON '{file_path}': {e}") from e
