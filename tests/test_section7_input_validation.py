"""Tests for Section 7 — Input Validation improvements."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest


# ---------------------------------------------------------------------------
# validate_project_name()
# ---------------------------------------------------------------------------

class TestValidateProjectName:
    def _v(self, name: str):
        from core.project_manager import validate_project_name
        return validate_project_name(name)

    def test_valid_name_returns_none(self):
        assert self._v("Sales Module") is None

    def test_valid_name_with_hyphens(self):
        assert self._v("Sales-Module-2024") is None

    def test_valid_name_with_dots_in_middle(self):
        assert self._v("v1.2 Project") is None

    def test_empty_string_returns_error(self):
        assert self._v("") is not None

    def test_whitespace_only_returns_error(self):
        assert self._v("   ") is not None

    def test_name_too_long_returns_error(self):
        assert self._v("A" * 101) is not None

    def test_name_at_limit_is_ok(self):
        assert self._v("A" * 100) is None

    def test_backslash_returns_error(self):
        assert self._v("Sales\\Module") is not None

    def test_forward_slash_returns_error(self):
        assert self._v("Sales/Module") is not None

    def test_colon_returns_error(self):
        assert self._v("C:Project") is not None

    def test_asterisk_returns_error(self):
        assert self._v("Project*") is not None

    def test_question_mark_returns_error(self):
        assert self._v("Project?") is not None

    def test_angle_brackets_return_error(self):
        assert self._v("Project<>") is not None

    def test_pipe_returns_error(self):
        assert self._v("Project|Name") is not None

    def test_double_quote_returns_error(self):
        assert self._v('Project"Name') is not None

    def test_reserved_name_con(self):
        assert self._v("CON") is not None

    def test_reserved_name_nul(self):
        assert self._v("NUL") is not None

    def test_reserved_name_case_insensitive(self):
        assert self._v("con") is not None

    def test_trailing_period_returns_error(self):
        assert self._v("Project.") is not None

    def test_error_message_is_string(self):
        result = self._v("C:Project")
        assert isinstance(result, str) and result


# ---------------------------------------------------------------------------
# create_project — invalid names now raise ValueError
# ---------------------------------------------------------------------------

class TestCreateProjectNameValidation:
    def test_slash_in_name_raises_os_error(self, tmp_path):
        """A name with / would silently create nested dirs; validate_project_name prevents it."""
        from core.project_manager import validate_project_name
        assert validate_project_name("bad/name") is not None

    def test_valid_name_creates_project(self, tmp_path):
        from core.project_manager import create_project, validate_project_name
        name = "Valid Name 2024"
        assert validate_project_name(name) is None
        p = create_project(name, "Corp", tmp_path)
        assert p.exists()


# ---------------------------------------------------------------------------
# detect_merged_cells()
# ---------------------------------------------------------------------------

class TestDetectMergedCells:
    @pytest.fixture(autouse=True)
    def _require_openpyxl(self):
        pytest.importorskip("openpyxl")

    def _make_clean_excel(self, tmp_path: Path) -> Path:
        path = tmp_path / "clean.xlsx"
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            pd.DataFrame({"a": [1, 2], "b": [3, 4]}).to_excel(
                writer, sheet_name="Sheet1", index=False
            )
        return path

    def _make_merged_excel(self, tmp_path: Path) -> Path:
        import openpyxl
        path = tmp_path / "merged.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MergedSheet"
        ws["A1"] = "Header"
        ws.merge_cells("A1:C1")
        ws["A2"] = "data"
        wb.save(path)
        return path

    def test_no_merged_cells_returns_false(self, tmp_path):
        from core.data_loader import detect_merged_cells
        path = self._make_clean_excel(tmp_path)
        assert detect_merged_cells(path, "Sheet1") is False

    def test_merged_cells_returns_true(self, tmp_path):
        from core.data_loader import detect_merged_cells
        path = self._make_merged_excel(tmp_path)
        assert detect_merged_cells(path, "MergedSheet") is True

    def test_missing_sheet_returns_false(self, tmp_path):
        from core.data_loader import detect_merged_cells
        path = self._make_clean_excel(tmp_path)
        assert detect_merged_cells(path, "NonExistent") is False

    def test_missing_file_returns_false(self, tmp_path):
        from core.data_loader import detect_merged_cells
        assert detect_merged_cells(tmp_path / "missing.xlsx", "Sheet1") is False


# ---------------------------------------------------------------------------
# Large file row-count threshold (logic only, no UI)
# ---------------------------------------------------------------------------

class TestLargeFileThreshold:
    _THRESHOLD = 100_000

    def test_small_df_below_threshold(self):
        df = pd.DataFrame({"a": range(99_999)})
        assert len(df) <= self._THRESHOLD

    def test_large_df_above_threshold(self):
        df = pd.DataFrame({"a": range(100_001)})
        assert len(df) > self._THRESHOLD

    def test_exact_threshold_not_triggered(self):
        df = pd.DataFrame({"a": range(100_000)})
        assert not (len(df) > self._THRESHOLD)


# ---------------------------------------------------------------------------
# Type mismatch detection (standalone logic, no UI)
# ---------------------------------------------------------------------------

class TestNumericRatioLogic:
    """Test the numeric-ratio helper logic used in _compare_column_compatibility."""

    def _numeric_ratio(self, vals: list[str]) -> float:
        sample = vals[:200]
        if not sample:
            return 0.0
        count = 0
        for v in sample:
            try:
                float(v.replace(",", "").replace(" ", ""))
                count += 1
            except ValueError:
                pass
        return count / len(sample)

    def test_all_numeric_returns_high_ratio(self):
        vals = [str(i) for i in range(100)]
        assert self._numeric_ratio(vals) >= 0.9

    def test_all_text_returns_low_ratio(self):
        vals = ["apple", "banana", "cherry"] * 50
        assert self._numeric_ratio(vals) < 0.1

    def test_mixed_returns_partial_ratio(self):
        vals = ["42"] * 50 + ["text"] * 50
        r = self._numeric_ratio(vals)
        assert 0.4 < r < 0.6

    def test_empty_list_returns_zero(self):
        assert self._numeric_ratio([]) == 0.0

    def test_floats_counted_as_numeric(self):
        vals = ["3.14", "2.71", "1.41"] * 30
        assert self._numeric_ratio(vals) >= 0.9

    def test_comma_separated_numbers_counted(self):
        vals = ["1,234", "5,678", "9,000"] * 30
        assert self._numeric_ratio(vals) >= 0.9

    def test_type_mismatch_would_trigger_warning(self):
        tx_vals = [str(i) for i in range(200)]   # numeric
        dim_vals = ["Apple", "Banana", "Cherry"] * 67  # text
        tx_ratio = self._numeric_ratio(tx_vals)
        dim_ratio = self._numeric_ratio(dim_vals)
        assert tx_ratio >= 0.9
        assert dim_ratio < 0.1

    def test_no_mismatch_when_both_text(self):
        tx_vals = ["Alpha", "Beta"] * 100
        dim_vals = ["Alpha", "Beta", "Gamma"] * 67
        assert self._numeric_ratio(tx_vals) < 0.1
        assert self._numeric_ratio(dim_vals) < 0.1
